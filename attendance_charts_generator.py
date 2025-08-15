import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import numpy as np
from io import BytesIO
# import os

def create_attendance_charts(excel_file):
    """
    Creates bar charts for Constituency_Monthly and Branch_Monthly sheets
    comparing monthly_attendance_avg vs target for each month.
    """
    
    # Read the Excel file
    xlsx = pd.ExcelFile(excel_file)
    
    # Read the data from both sheets
    constituency_df = pd.read_excel(xlsx, 'Constituency_Monthly')
    branch_df = pd.read_excel(xlsx, 'Branch_Monthly')
    
    # Create a month order for proper sorting
    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 
                   'July', 'August', 'September', 'October', 'November', 'December']
    
    # Convert month names to categorical for proper ordering
    constituency_df['Month'] = pd.Categorical(constituency_df['Month'], categories=month_order, ordered=True)
    branch_df['Month'] = pd.Categorical(branch_df['Month'], categories=month_order, ordered=True)
    
    # Load the workbook to add new sheets
    wb = load_workbook(excel_file)
    
    # Create Constituency Monthly Charts
    print("Creating Constituency Monthly Charts...")
    create_constituency_charts(constituency_df, wb)
    
    # Create Branch Monthly Charts
    print("Creating Branch Monthly Charts...")
    create_branch_charts(branch_df, wb)
    
    # Save the workbook with new charts
    output_file = excel_file.replace('.xlsx', '_with_charts.xlsx')
    wb.save(output_file)
    print(f"\nCharts saved to: {output_file}")
    
    return output_file


def create_constituency_charts(df, workbook):
    """Creates bar charts for each constituency and adds them to a new sheet."""
    
    # Create or get the charts sheet
    if 'Constituency_Monthly_Charts' in workbook.sheetnames:
        ws = workbook['Constituency_Monthly_Charts']
    else:
        ws = workbook.create_sheet('Constituency_Monthly_Charts')
    
    # Get unique constituencies
    constituencies = df['Constituency'].unique()
    
    # Create figure with subplots
    n_constituencies = len(constituencies)
    n_cols = 2
    n_rows = (n_constituencies + n_cols - 1) // n_cols
    
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(15, 5 * n_rows))
    fig.suptitle('Monthly Average Attendance vs Target by Constituency', fontsize=16, fontweight='bold')
    
    # Flatten axes array for easier iteration
    if n_rows == 1:
        axes = axes.reshape(1, -1)
    
    # Plot each constituency
    for idx, constituency in enumerate(constituencies):
        row = idx // n_cols
        col = idx % n_cols
        ax = axes[row, col]
        
        # Filter data for this constituency
        const_data = df[df['Constituency'] == constituency].sort_values('Month')
        
        # Create bar positions
        months = const_data['Month'].astype(str)
        x = np.arange(len(months))
        width = 0.35
        
        # Create bars
        bars1 = ax.bar(x - width/2, const_data['Monthly_Attendance_Avg'], 
                       width, label='Attendance', color='skyblue')
        bars2 = ax.bar(x + width/2, const_data['Target'], 
                       width, label='Target', color='coral')
        
        # Customize the subplot
        ax.set_xlabel('Month', fontweight='bold')
        ax.set_ylabel('Attendance', fontweight='bold')
        ax.set_title(f'{constituency}', fontweight='bold', fontsize=12)
        ax.set_xticks(x)
        ax.set_xticklabels(months, rotation=45, ha='right')
        # ax.legend()
        ax.grid(True, alpha=0.3, axis='y')
        
        # Add value labels on bars
        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                ax.annotate(f'{height:.1f}',
                           xy=(bar.get_x() + bar.get_width() / 2, height),
                           xytext=(0, 3),
                           textcoords="offset points",
                           ha='center', va='bottom',
                           fontsize=8)
    
    # Hide empty subplots
    for idx in range(n_constituencies, n_rows * n_cols):
        row = idx // n_cols
        col = idx % n_cols
        axes[row, col].set_visible(False)
    
    # Collect unique handles and labels
    handles, labels = [], []
    for ax in fig.axes:
        for handle, label in zip(*ax.get_legend_handles_labels()):
            if label not in labels:  # Add only unique labels and their corresponding handles
                handles.append(handle)
                labels.append(label)
    
    # Create a single legend for the entire figure
    fig.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, 1.05), ncol=len(labels))
    
    plt.tight_layout()
    
    # Save to BytesIO and add to Excel
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
    plt.close()
    
    img_buffer.seek(0)
    img = Image(img_buffer)
    ws.add_image(img, 'A1')


def create_branch_charts(df, workbook):
    """Creates bar charts for each branch grouped by constituency."""
    
    # Create or get the charts sheet
    if 'Branch_Monthly_Charts' in workbook.sheetnames:
        ws = workbook['Branch_Monthly_Charts']
    else:
        ws = workbook.create_sheet('Branch_Monthly_Charts')
    
    # Get unique constituencies for grouping
    constituencies = df['Constituency'].unique()
    
    row_position = 1
    
    for const_idx, constituency in enumerate(constituencies):
        # Filter branches for this constituency
        const_branches = df[df['Constituency'] == constituency]
        branches = const_branches['Branch'].unique()
        
        # Create figure for this constituency's branches
        n_branches = len(branches)
        n_cols = min(3, n_branches)  # Max 3 columns per constituency group
        n_rows = (n_branches + n_cols - 1) // n_cols
        
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(15, 5 * n_rows))
        fig.suptitle(f'{constituency} - Monthly Average Attendance vs Target by Branch', 
                     fontsize=14, fontweight='bold')
        
        # Handle single subplot case
        if n_branches == 1:
            axes = np.array([[axes]])
        elif n_rows == 1:
            axes = axes.reshape(1, -1)
        elif n_cols == 1:
            axes = axes.reshape(-1, 1)
        
        # Plot each branch
        for idx, branch in enumerate(branches):
            row = idx // n_cols
            col = idx % n_cols
            ax = axes[row, col]
            
            # Filter data for this branch
            branch_data = const_branches[const_branches['Branch'] == branch].sort_values('Month')
            
            # Create bar positions
            months = branch_data['Month'].astype(str)
            x = np.arange(len(months))
            width = 0.35
            
            # Create bars
            bars1 = ax.bar(x - width/2, branch_data['Monthly_Attendance_Avg'], 
                           width, label='Attendance', color='lightgreen')
            bars2 = ax.bar(x + width/2, branch_data['Target'], 
                           width, label='Target', color='salmon')
            
            # Customize the subplot
            ax.set_xlabel('Month', fontweight='bold')
            ax.set_ylabel('Attendance', fontweight='bold')
            ax.set_title(f'{branch}', fontweight='bold', fontsize=11)
            ax.set_xticks(x)
            ax.set_xticklabels(months, rotation=45, ha='right')
            # ax.legend()
            ax.grid(True, alpha=0.3, axis='y')
            
            # Add value labels on bars
            for bars in [bars1, bars2]:
                for bar in bars:
                    height = bar.get_height()
                    ax.annotate(f'{height:.1f}',
                               xy=(bar.get_x() + bar.get_width() / 2, height),
                               xytext=(0, 3),
                               textcoords="offset points",
                               ha='center', va='bottom',
                               fontsize=8)
        
        # Hide empty subplots
        for idx in range(n_branches, n_rows * n_cols):
            row = idx // n_cols
            col = idx % n_cols
            axes[row, col].set_visible(False)
        
        # Collect unique handles and labels
        handles, labels = [], []
        for ax in fig.axes:
            for handle, label in zip(*ax.get_legend_handles_labels()):
                if label not in labels:  # Add only unique labels and their corresponding handles
                    handles.append(handle)
                    labels.append(label)
        
        # Create a single legend for the entire figure
        fig.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, 1.05), ncol=len(labels))
        
        plt.tight_layout()
        
        # Save to BytesIO and add to Excel
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        plt.close()
        
        img_buffer.seek(0)
        img = Image(img_buffer)
        
        # Calculate position for this constituency's chart
        cell_position = f'A{row_position}'
        ws.add_image(img, cell_position)
        
        # Update row position for next constituency (estimate ~40 rows per chart)
        row_position += 40


# Example usage
if __name__ == "__main__":
    # Replace with your Excel file path
    excel_file = "monthly_attendance_analysis_results.xlsx"
    
    # Create the charts
    output_file = create_attendance_charts(excel_file)
    print(f"Process completed! Charts have been added to: {output_file}")
